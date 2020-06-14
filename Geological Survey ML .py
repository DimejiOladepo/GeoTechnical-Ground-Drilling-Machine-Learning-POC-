#!/usr/bin/env python
# coding: utf-8

# In[1]:


# Import time library and start timing the program
import time
start_time = time.time()
#Import frameworks used by program
import warnings
warnings.filterwarnings('ignore')
import re
import openpyxl
import numpy as np
import pandas as pd 
import os
from sklearn.model_selection import train_test_split, GridSearchCV, cross_val_score
from sklearn import metrics
from lightgbm import LGBMClassifier


# In[ ]:


def flattenNestedList(nestedList):
    ''' Helper function which converts a nested list to a flat list '''
    flatList = []
    # Iterate over all the elements in given list
    for elem in nestedList:
        # Check if type of element is list
        if isinstance(elem, list):
            # Extend the flat list by adding contents of this element (list)
            flatList.extend(flattenNestedList(elem))
        else:
            # Append the element to the list
            flatList.append(elem)    
 
    return flatList


# In[ ]:


def datapairlist(folder):
    
    """This function creates a one to one list of lists containing pairs of 
Cone Penetration Test SND files and their corresponding Laboratory Test results in 
Workbook per sub-folder 
    """
    file_pair_list = list()
    #Listing all subfolders in the GEOML-1 folder
    geoML_subfolders = [ sub.path for sub in os.scandir(folder) if sub.is_dir() ]

    for subfolder in geoML_subfolders:
        #Creating list of directory path to laboratory report workbook files  
        workbook_dir_list = list()
        for root, dirs, files in os.walk(subfolder):
            for file in files:
                if file.endswith(".xlsm") or file.endswith(".xlsx"):
                    workbook_dir_list.append(os.path.join(root,file))
        #Extracting list of workbook id number 
        workbook_id = list()
        for _id in workbook_dir_list:
            workbook_id.append(_id[:-5].split()[-1])
            
        #Create dictionary of workbook id to workbook directory path
        id_dictionary = dict(zip(workbook_id, workbook_dir_list))

        #Create list of SND files in subfolders
        snd_name_list = list()
        for root, dirs, files in os.walk(subfolder):
            for file in files:
                if file.endswith(".SND"):
                    snd_name_list.append(file)
        
        #Match workbook id with SNDs 
        snd_match = list()
        for element in snd_name_list:
            if element[:-4] in workbook_id:
                snd_match.append(os.path.join(root, element))

        #Add matching Workbook to SNDs to list         
        workbook_match = list()
        for elements in snd_match:
            identifier = elements[:-4].split("/")
            workbook_match.append(id_dictionary[identifier[-1]])

        #Create list of list of folder lists containing Workbook-SND pairs matched by id 
        file_pairs = list(zip(workbook_match, snd_match))
        file_pair_list.append(file_pairs)
        
    #Filter out empty lists
    file_pair_list_filtered = list(filter(None, file_pair_list))
    
    #Creating list of list from file_pairs
    pair_lst = list()
    for level_2 in file_pair_list_filtered:
        for level_1 in level_2:
            pair_lst.append(list(level_1))
    
    return(pair_lst)


# In[ ]:


#Dictionary containing Soil type translation from Norwegian to English
soil_type = {'LEIRE': 'CLAY', 'KVIKKLEIRE': 'QUICK CLAY', 'TØRRSKORPELEIRE': 'WEATHERED CLAY', 
             'SILT': 'SILT', 'TØRRSKORPESILT': 'WEATHERED SILT', 'SAND': 'SAND', 
             'GRUS': 'GRAVEL', 'TORV': 'PEAT', 'GYTJE': 'GYTJA', 
             'ORG. MATR.': 'ORG. MAT.', 'MATJORD': 'TOPSOIL', 'DY': 'DY',
             'MATERIALE': 'MATERIAL', 'FYLLMASSE': 'FILL SOIL'}

#List for incompatible pairs which are excluded
incompatible_pairs = []

#Creating empty dataframe with columns for housing extracted data
big_merge_df = pd.DataFrame(columns = ["Drill Depth (m)", "Soil type", "X-coordinate", "Y-coordinate", 
                                       "Height Above Sea Level", "Drill Pressure (kN)","Flushing Pressure (kN)"])

for pair in datapairlist("GEOML-1"):
    try:
        # read SNDs in datapairlist as comma separated values
        snd_table = pd.read_csv(pair[1], delimiter = "\r\n", header = None, sep = " ", names = "v")
        snd_table_lst = []
        # read values in 15th row and below
        for row in snd_table[15:].values.tolist():
            for column in row:
                snd_table_lst.append(column.split())

        # create dataframe from snd table list          
        df = pd.DataFrame(snd_table_lst)
        # use only four columns in dataframe and add column names "Drill Depth (m)", "Drill Pressure (kN)",
        # "Torque", "Flushing Pressure (kN)".
        snd_table_df = df.iloc[:, :4]
        snd_table_df.columns = ["Drill Depth (m)", "Drill Pressure (kN)", "Torque", "Flushing Pressure (kN)"]

        cleaned_table = snd_table_df
        
        # Extract first two values and transpose arrangement to two different columns 
        position = snd_table[:2].T
        # Name columns "X-coordinate" and "Y-coordinate"
        position.columns = ["X-coordinate", "Y-coordinate"]
        # Populate position column values with as many values in snd_table_df
        position = position.append([position]*(len(cleaned_table)-1), ignore_index=True)

        #Extracting height above sea level value and deducting value per drill step
        height = float(snd_table['v'][2])
        height_box = []
        drill_step = round(float(df[0][1]) - float(df[0][0]),4)
        for depth in range(len(cleaned_table)):
            height -= drill_step
            height_box.append(round(height,3))
            
        #Create dataframe for changing depth
        height_df = pd.DataFrame(height_box, columns = ['Height Above Sea Level'])
        
        #Merging position, height_df and cleaned_table dataframes 
        snd_df = pd.concat([position, height_df, cleaned_table], axis = 1)

        #Converting all columns in merged dataframe to numeric data type 
        for i in range(0, len(snd_df.columns)):
            snd_df.iloc[:,i] = pd.to_numeric(snd_df.iloc[:,i], errors='ignore')
        #Removing rows with empty Torque values 
        snd_df = snd_df[snd_df.Torque.notnull()]
        #Converting Drill Depth column to float 
        snd_df["Drill Depth (m)"] = snd_df["Drill Depth (m)"].astype(float)
        
        #Read workbook in datapairlist
        book = openpyxl.load_workbook(pair[0], data_only = True)
        sheetname_list = []
        #Specifying search pattern for only spreadsheets named with integer numbers
        pattern = '^[0-9]*$'
        #Search through sheets and only add spreadsheet names named with integer numbers
        for sheetname in book.sheetnames:
            if re.match(pattern, sheetname):
                sheetname_list.append(sheetname)
        workbook_depth_list = []        
        workbook_material_list = []
        # Go through spreadsheet and extract soil depth tested for
        for sheet in sheetname_list:
            sheet_depth_list = []
            for value in book[sheet].iter_rows(min_row=8,
                                      max_row=24,
                                      min_col=4,
                                      max_col=4,
                                      values_only=True):
                if value[0] is None:
                    pass
                else:
                    sheet_depth_list.append(value[0]) 
            #If values exist in selected spreadsheet, find the min and add drill step value till max is reached 
            #This captures the values in-between the max and min in the sheet
            try:
                min_depth_value = min(sheet_depth_list)
                max_depth_value = max(sheet_depth_list)

                sheet_depth_range_list = [min_depth_value]

                while max_depth_value > round(min_depth_value,3):
                    min_depth_value += drill_step
                    sheet_depth_range_list.append(round(min_depth_value,3))
                workbook_depth_list.append(sheet_depth_range_list)
                
            #Extract soil type for spreadsheet and duplicate values for length of spreadsheet depth
                for val in book[sheet].iter_rows(min_row=24,
                                          max_row=24,
                                          min_col=15,
                                          max_col=15,
                                          values_only = True):
                    soil_list = [soil_type[val[0]] for material in range(len(sheet_depth_range_list))]
                workbook_material_list.append(soil_list)
            except:
                continue
            #Flattening nested lists to singular list 
            workbook_depth_list = flattenNestedList(workbook_depth_list)
            workbook_material_list = flattenNestedList(workbook_material_list)
            
            workbook_df = pd.DataFrame(list(zip(workbook_depth_list, workbook_material_list)),
                                       columns =['Drill Depth (m)', 'Soil type']) 
            
            #Converting all workbook columns to numeric type
            for i in range(0, len(workbook_df.columns)):
                workbook_df.iloc[:,i] = pd.to_numeric(workbook_df.iloc[:,i], errors='ignore')
                
        #Merging SND and Workbook Dataframes into one
        merged_df = pd.merge(left=workbook_df, right=snd_df, how='left', left_on='Drill Depth (m)',
                             right_on='Drill Depth (m)')
        
        #Adding all merged dataframes to one Dataframe
        big_merge_df = big_merge_df.append(merged_df, ignore_index = True, sort = False)

    except:
        #Adding all incompatible pairs to incompatible_pairs list
        incompatible_pairs.append(pair)
        continue


# In[ ]:


#Exporting big_merge_df to csv file to be saved locally at the same file location
big_merge_df.to_csv('data_extract.csv')


# In[ ]:


# Model building and Feature Engineering 
# Reading exported dataset 
dataset = pd.read_csv(r"data_extract.csv", index_col = 0)
# Dropping rows containing emoty cells 
dataset = dataset.dropna(axis=0, how='any', thresh=None, subset=None, inplace=False)

# Splitting dataset into independent and dependent sets where soil type is the dependent set
y = dataset["Soil type"]
X = dataset.drop("Soil type", axis = 1)

#Splitting the sets into train and test sets with test set size set to 30% of the data 
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size = 0.3, random_state = 50) 

#Using LightGBM model to learn from the train set.
model = LGBMClassifier()
model.fit(X_train, y_train)

#Testing some hyperparameters for Lightgbm model to improve model performance 
gridParams = {
    'learning_rate': [0.1, 0.05],
    'n_estimators': [40, 200, 400],
    'num_leaves': [20, 40],
    'boosting_type' : ['gbdt'],
    'objective' : ['multiclass'],
    'random_state' : [42], 
    'colsample_bytree' : [0.8, 1],
    'subsample' : [0.75,1],
    'reg_alpha' : [1,0.5],
    'reg_lambda' : [1,0.5],
    }
#Using gridsearch to test through specified model hyperparameters and determine best values
grid = GridSearchCV(model, gridParams,
                    verbose=0,
                    cv=4,
                    n_jobs=2)
#fitting gridsearch on train data
grid.fit(X_train, y_train)
print(grid.best_params_)
print("")


# In[ ]:


#Specifying model hyperparameters from gridsearch 
gbm = LGBMClassifier(boosting_type = 'gbdt', colsample_bytree = 0.8, learning_rate = 0.1, 
                     n_estimators = 400, num_leaves = 20, objective = 'multiclass', 
                     random_state = 42, reg_alpha = 0.5, reg_lambda = 0.5, subsample = 0.75)
#Refitting model 
gbm.fit(X_train, y_train)
#Using model to predict test dependent variable
Y_sum = gbm.predict(X_test)

#Model Accuracy
accuracy = metrics.accuracy_score(y_test, Y_sum)
print('Accuracy: %f' % accuracy)
#Model Precision
precision = metrics.precision_score(y_test, Y_sum, average = 'macro')
print('Precision: %f' % precision)
#Model Recall
recall = metrics.recall_score(y_test, Y_sum, average = 'macro')
print('Recall: %f' % recall)
#Model F1 score 
f1 = metrics.f1_score(y_test, Y_sum, average = 'macro')
print('F1 score: %f' % f1)

print("")
#Model Cross validation score 
print('Cross-Validation Score: %f' % np.mean(cross_val_score(gbm, X, y, cv=10)))

#Stop program time 
print("--- %s seconds ---" % (time.time() - start_time))

